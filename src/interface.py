import sys

is_frozen = getattr(sys, 'frozen', False)
frozen_temp_path = getattr(sys, '_MEIPASS', '')

import os

# This is needed to find resources when using pyinstaller
if is_frozen:
    basedir = frozen_temp_path
else:
    basedir = os.path.dirname(os.path.abspath(__file__))
resource_dir = os.path.join(basedir, 'resources')



import gi
gi.require_version("Gtk", "3.0")
from gi.repository import Gtk, Gdk, GObject, GLib



from GridWindow import GridWindow
from About import AboutWindow
from CoilsListBox import CoilsListBox
from CoilListRow import CoilListRow
from Presets import HelmholtzCoilPreset
from Presets import MaxwellCoilPreset
from Presets import WangCoilPreset
from Presets import TetraCoilPreset
from Presets import LeeWhitingCoilPreset
from Presets import RandomCoilPreset
from coil import Coil, CreateCoil
from Simulation import Simulation
from Results import Results
from ErrorMessage import ErrorMessage
import random
import numpy

import openpyxl


class InputWindow():
    def __init__(self, glade_file):
        self.glade_file = glade_file
        self.builder = Gtk.Builder()
        self.builder.add_from_file(glade_file)
        
        self.window = self.builder.get_object("wndInput")
        
        self.btnHelmholtzConfig = self.builder.get_object("btnHelmholtzConfig")
        self.btnMaxwellConfig = self.builder.get_object("btnMaxwellConfig")
        self.btnWangConfig = self.builder.get_object("btnWangConfig")
        self.btnTetracoilConfig = self.builder.get_object("btnTetracoilConfig")
        self.btnLeeConfig = self.builder.get_object("btnLeeConfig")
        self.btnRandomConfig = self.builder.get_object("btnRandomConfig")
        
        self.scrListBox = self.builder.get_object("scrListBox")
        self.btnSimulate = self.builder.get_object("btnSimulate")
        self.chbAutoGrid = self.builder.get_object("chbAutoGrid")
        self.menuColorMap = self.builder.get_object("menuColorMap")
        self.treeData = self.builder.get_object("treeData")
        self.btnLoadParams = self.builder.get_object("btnLoadParams")
        self.btnLoadResults = self.builder.get_object("btnLoadResults")
        self.btnNew = self.builder.get_object("btnNew")
        self.btnQuit = self.builder.get_object("btnQuit")
        self.btnAbout = self.builder.get_object("btnAbout")
        
        self.listBox = CoilsListBox(self.btnSimulate)
        self.scrListBox.add_with_viewport(self.listBox)

        self.window.connect("destroy", Gtk.main_quit)

        self.btnHelmholtzConfig.connect("activate", self.on_helmholtz_config)
        self.btnMaxwellConfig.connect("activate", self.on_maxwell_config)
        self.btnWangConfig.connect("activate", self.on_wang_config)
        self.btnTetracoilConfig.connect("activate", self.on_tetracoil_config)
        self.btnLeeConfig.connect("activate", self.on_lee_config)
        self.btnRandomConfig.connect("activate", self.on_random_config)

        self.btnSimulate.connect("clicked", self.on_simulate)
        self.chbAutoGrid.connect("toggled", self.on_auto_grid)
        self.btnLoadParams.connect("activate", self.on_import_params)
        self.btnLoadResults.connect("activate", self.on_import_results)
        self.btnQuit.connect("activate", Gtk.main_quit)
        self.btnNew.connect("activate", self.listBox.remove_all_coils)
        self.btnAbout.connect("activate", lambda _: AboutWindow(self.window))



        self.auto_grid = self.chbAutoGrid.get_active()
        self.coils = []
        self.z_min = 0.0
        self.z_max = 0.0
        self.z_points = 0
        self.y_min = 0.0
        self.y_max = 0.0
        self.y_points = 0

        self.window.show_all()
        self.window.maximize()
        self.listBox.create_coil_row(None)
        
        Gtk.main()



    def on_helmholtz_config(self, widget):
        self.listBox.update(HelmholtzCoilPreset())

    def on_maxwell_config(self, widget):
        self.listBox.update(MaxwellCoilPreset())

    def on_wang_config(self, widget):
        self.listBox.update(WangCoilPreset())

    def on_tetracoil_config(self, widget):
        self.listBox.update(TetraCoilPreset())

    def on_lee_config(self, widget):
        self.listBox.update(LeeWhitingCoilPreset())

    def on_random_config(self, widget):
        self.listBox.update(RandomCoilPreset(random.randint(2, 10)))


    def on_auto_grid(self, check):
        self.auto_grid = check.get_active()
        # print(self.auto_grid)

    def compute_grid(self):
        if len(self.coils) > 0:
            z_arr = [coil.pos_z for coil in self.coils]
            self.z_min = min(z_arr)
            self.z_max = max(z_arr)

            radius_arr = [coil.radius for coil in self.coils]
            self.y_min = -max(radius_arr)
            self.y_max = max(radius_arr)

            if self.z_min == self.z_max:
                self.z_min = self.z_min - self.y_max
                self.z_max = self.z_max + self.y_max

            PMAX = 100
            if abs(self.z_max - self.z_min) > abs(self.y_max - self.y_min):
                self.z_points = PMAX
                self.y_points = int(abs(self.y_max - self.y_min) * PMAX / abs(self.z_max - self.z_min))
            else:
                self.y_points = PMAX
                self.z_points = int(abs(self.z_max - self.z_min) * PMAX / abs(self.y_max - self.y_min))


    def validate_values(self, coil_row):
        values = coil_row.get_values()
        radius = values["radius"]
        turns = values["turns"]
        current = values["current"]
        position = values["position"]
        # print(radius, turns, current, position)
        
        if not (radius and radius > 0.0):
            ErrorMessage(self.window, "Invalid input parameters", "Radius must be a positive real.")
            return False

        if not (turns and turns > 0):
            ErrorMessage(self.window, "Invalid input parameters", "Number of turns must be a positive integer.")
            return False

        if not (current and abs(current) < 150):
            ErrorMessage(self.window, "Invalid input parameters", "Electric current must be a real between -150 and 150.")
            return False

        if (isinstance(position, bool) and not position):
            ErrorMessage(self.window, "Invalid input parameters", "Position must be a real number.")
            return False

        return True

    def collect_coils_values(self):
        self.coils = []
        for row in list(self.listBox)[:-1]:
            coil_row, = row.get_children()
            val = self.validate_values(coil_row)
            if not val:
                return False
            
            coil = CreateCoil(**coil_row.get_values())
            self.coils.append(coil)
        return True


    def on_simulate(self, widget):
        flag = self.collect_coils_values()
        if not flag:
            return

        if len(self.coils) == 0:
            ErrorMessage(self.window, "Invalid input parameters", "At least one coil must be added.")
            return

        self.compute_grid()
        ready = True
        if not self.auto_grid:
            ready = self.insert_grid_manually()

        if ready:
            # print("lets go")
            self.simulation = Simulation(self, self.coils,
                self.z_min, self.z_max, self.z_points,
                self.y_min, self.y_max, self.y_points)
            self.simulation.simulate()


    def isNumeric(self, val, func=float):
        try:
            func(val)
            return True
        except Exception as e:
            return False

    def insert_grid_manually(self):
        initial_grid = {
            "z_min": self.z_min,
            "z_max": self.z_max,
            "z_points": self.z_points,
            "y_min": self.y_min,
            "y_max": self.y_max,
            "y_points": self.y_points,
        }

        
        dialog = GridWindow(self.window, resource_dir + "/grid.glade", initial_grid)
        
        response = dialog.window.run()

        if response == Gtk.ResponseType.OK:
            self.z_min = float(dialog.txtMinZ.get_text()) if self.isNumeric(dialog.txtMinZ.get_text()) else False
            self.z_max = float(dialog.txtMaxZ.get_text()) if self.isNumeric(dialog.txtMaxZ.get_text()) else False
            self.z_points = int(dialog.txtPointsZ.get_text()) if self.isNumeric(dialog.txtPointsZ.get_text(), int) else False
            self.y_min = float(dialog.txtMinY.get_text()) if self.isNumeric(dialog.txtMinY.get_text()) else False
            self.y_max = float(dialog.txtMaxY.get_text()) if self.isNumeric(dialog.txtMaxY.get_text()) else False
            self.y_points = int(dialog.txtPointsY.get_text()) if self.isNumeric(dialog.txtPointsY.get_text(), int) else False
            dialog.window.destroy()

            if (isinstance(self.z_min, bool) or 
                isinstance(self.z_max, bool) or 
                isinstance(self.y_min, bool) or 
                isinstance(self.y_max, bool)):
                if not (self.z_min and self.z_max and self.y_min and self.y_max):
                    ErrorMessage(self.window, "Invalid input parameters", "Simulation limits must be real numbers.")
                    return False

            if not (self.z_min < self.z_max and self.y_min < self.y_max):
                ErrorMessage(self.window, "Invalid input parameters", "Min. value must be lower than Max. value.")
                return False

            if not (self.z_points and self.y_points and self.z_points > 0 and self.y_points > 0):
                ErrorMessage(self.window, "Invalid input parameters", "Number of simulation points must be greater than 0.")
                return False

            return True

        dialog.window.destroy()
        return False

    def on_import_params(self, widget):
        dialog = Gtk.FileChooserDialog("Please choose a file", self.window,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        filters = Gtk.FileFilter()
        filters.set_name("Excel files")
        filters.add_pattern("*.*.csv")
        filters.add_pattern("*.xlsx")
        filters.add_pattern("*.XLSX")
        dialog.add_filter(filters)

        response = dialog.run()
        
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()

            wb = openpyxl.load_workbook(filename)
            wInput = wb["Simulation parameters"]
            wCoils = wb['Input parameters']

            self.z_min = wInput.cell(row=1 + 0, column=1 + 1).value
            self.z_max = wInput.cell(row=1 + 1, column=1 + 1).value
            self.z_points = int(wInput.cell(row=1 + 2, column=1 + 1).value)
            self.y_min = wInput.cell(row=1 + 3, column=1 + 1).value
            self.y_max = wInput.cell(row=1 + 4, column=1 + 1).value
            self.y_points = int(wInput.cell(row=1 + 5, column=1 + 1).value)

            coils = []
            for i in range(wCoils.max_row - 1):
                radius = wCoils.cell(row=1 + i + 1, column=1 + 0).value
                turns = int(wCoils.cell(row=1 + i + 1, column=1 + 1).value)
                current = wCoils.cell(row=1 + i + 1, column=1 + 2).value
                position = wCoils.cell(row=1 + i + 1, column=1 + 3).value
                coils.append(CreateCoil("Circular", radius, turns, current, position))

            coil_rows = []
            for coil in coils:
                coil_row = CoilListRow()
                coil_row.set_values(
                    radius=coil.radius,
                    turns=coil.num_turns,
                    current=coil.I, position=coil.pos_z)
                coil_rows.append(coil_row)
            self.listBox.update(coil_rows)

        elif response == Gtk.ResponseType.CANCEL:
            pass
            # print("Cancel clicked")

        dialog.destroy()

    def on_import_results(self, widget):
        dialog = Gtk.FileChooserDialog("Please choose a file", self.window,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        filters = Gtk.FileFilter()
        filters.set_name("Excel files")
        filters.add_pattern("*.*.csv")
        filters.add_pattern("*.xlsx")
        filters.add_pattern("*.XLSX")
        dialog.add_filter(filters)

        response = dialog.run()
        
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()


            wb = openpyxl.load_workbook(filename)
            wInput = wb["Simulation parameters"]
            wCoils = wb['Input parameters']
            wBy = wb['B y']
            wBz = wb['B z']
            wBnorm = wb['B norm']

            z_min = wInput.cell(row=1 +0, column=1 + 1).value
            z_max = wInput.cell(row=1 +1, column=1 + 1).value
            z_points = int(wInput.cell(row=1 +2, column=1 + 1).value)
            y_min = wInput.cell(row=1 +3, column=1 + 1).value
            y_max = wInput.cell(row=1 +4, column=1 + 1).value
            y_points = int(wInput.cell(row=1 +5, column=1 + 1).value)

            coils = []
            for i in range(wCoils.max_row - 1):
                radius = wCoils.cell(row=1 + i + 1, column=1 + 0).value
                turns = int(wCoils.cell(row=1 + i + 1, column=1 + 1).value)
                current = wCoils.cell(row=1 + i + 1, column=1 + 2).value
                position = wCoils.cell(row=1 + i + 1, column=1 + 3).value
                coils.append(CreateCoil("Circular", radius, turns, current, position))

            z_arr = []
            for i in range(wBz.max_column - 1):
                z_arr.append(wBz.cell(row=1 + 0, column=1 + i + 1).value)

            y_arr = []
            for i in range(wBz.max_row - 1):
                y_arr.append(wBz.cell(row=1 + i + 1, column=1 + 0).value)

            Bz_grid = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            Brho_grid = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            norm = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            for i in range(len(z_arr) - 1):
                for j in range(len(y_arr) - 1):
                    Bz_grid[i, j] = wBz.cell(row=1 + j + 1, column=1 + i + 1).value
                    Brho_grid[i, j] = wBy.cell(row=1 + j + 1, column=1 + i + 1).value
                    norm[i, j] = wBnorm.cell(row=1 + j + 1, column=1 + i + 1).value

            self.simulation = Simulation(self, coils,
                            z_min, z_max, 1,
                            y_min, y_max, 1)
            self.simulation.set_data(coils, z_min, z_max, z_points, y_min, y_max, y_points,
                 z_arr, y_arr, Bz_grid, Brho_grid, norm)

            results = Results(self, self.simulation)
            results.load_simulation()

            coil_rows = []
            for coil in coils:
                coil_row = CoilListRow()
                coil_row.set_values(
                    radius=coil.radius,
                    turns=coil.num_turns,
                    current=coil.I, position=coil.pos_z)
                coil_rows.append(coil_row)
            self.listBox.update(coil_rows)
            self.window.hide()

        elif response == Gtk.ResponseType.CANCEL:
            pass
            # print("Cancel clicked")

        dialog.destroy()

GObject.threads_init()
window = InputWindow(resource_dir + "/input.glade")
